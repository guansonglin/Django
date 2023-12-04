
from django.shortcuts import render,redirect,HttpResponse
from app01 import models
from app01.utils.pagination import Pagination


def depart_list(request):
    ''' 部门列表 '''

    queryset = models.Department.objects.all()
    page_object = Pagination(request,queryset)
    context = {
        "queryset" : page_object.page_queryset,
        "page_string" : page_object.html(),
        "error" : page_object.error
        }

    return render(request , "depart_list.html" , context)


def depart_add(request):
    ''' 部门添加 '''

    if request.method == 'GET':
        return render(request , "depart_add.html")
    else:
        title = request.POST.get('title')
        models.Department.objects.create(title = title)

        return redirect('/depart/list/')


def depart_delete(request):
    ''' 部门删除 '''
    
    ider = request.GET.get('ider')
    models.Department.objects.filter(id = ider).delete()
 
    return redirect("/depart/list/")


def depart_edit(request,ider):
    ''' 部门编辑 '''

    if request.method == "GET":
        old_data = models.Department.objects.filter(id = ider).first()
        # print(old_data.id,old_data.title)
        return render(request , "depart_edit.html" ,{"old_data" : old_data})
        
    else:

        new_title = request.POST.get("title")
        models.Department.objects.filter(id = ider).update(title = new_title)
        return redirect("/depart/list/")


def depart_multi(request):
    ''' 批量上传（Excel文件） '''
    from openpyxl import load_workbook


    file_object = request.FILES.get("exc")
    # print(file_object)
    # print(type(file_object)) # 获取的是一个文件对象 <class 'django.core.files.uploadedfile.InMemoryUploadedFile'>

    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # cell = sheet.cell(1,1) #获取第一行第一列
    # print(cell.value)

    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        
        exists = models.Department.objects.filter(title= text).exists()
        if not exists:
            models.Department.objects.create(title = text)

    
    return redirect('/depart/list/')

